#!/usr/bin/python3
# -*- coding: utf-8 -*
import tkinter
from tkinter import Tk, Frame, Menu, ttk
from tkinter import *

from openpyxl import load_workbook

class Example(tkinter.Frame):
    def __init__(self):
        """ initialise UI with menu bar upon opening """
        super().__init__()
        self.QuestionRowsToAdd = [] # accumulate rows of questions to add to the survey sheet
        self.ChoiceRowsToAdd = [] # accumulate list of choice options to add to choices sheet

        self.initUI()

    def initUI(self):
        self.master.title("Survey")  # set window title

        menubar = tkinter.Menu(self.master)  # add menu bar
        self.master.config(menu=menubar)

        # under "File"
        fileMenu = tkinter.Menu(menubar, tearoff=0)
        fileMenu.add_command(label="New Question",
                             command=lambda:self.displayAddNewQuestion(self.master))  # form to add new question
        fileMenu.add_command(label="Save", command=self.save)  # save changes
        fileMenu.add_separator()
        fileMenu.add_command(label="Quit", command=self.onExit)  # close window
        menubar.add_cascade(label="File", menu=fileMenu)

        # under "View"
        viewMenu = tkinter.Menu(menubar, tearoff=0)
        viewMenu.add_command(label="Question Additions", command=self.displayViewQuestionAdditions)
        viewMenu.add_command(label="Choice Additions", command=self.displayViewChoiceAdditions)
        menubar.add_cascade(label="View", menu=viewMenu)

        # under "Help"
        helpMenu = tkinter.Menu(menubar, tearoff=0)
        helpMenu.add_command(label="About...", command=self.displayHelpMenu)  # display help popup
        menubar.add_cascade(label="Help", menu=helpMenu)

    def displayViewQuestionAdditions(self):
        """ Display pop up menu showing current question additions """
        viewPopup = Tk()  # pop-up
        viewPopup.wm_title("Question Additions")

        # TODO
        ttk.Label(viewPopup, text="add some explanation here").pack(side="top", fill="x", pady=10)
        ttk.Label(viewPopup, text="<id>: <label>").pack(side="top", fill="x", pady=20)

        # question = [type, name, label, ...]
        for question in self.QuestionRowsToAdd:
            currentQ = question[1] + ": " + question[2]
            ttk.Label(viewPopup, text=currentQ).pack(side="top", fill="x", pady=10)

    def displayViewChoiceAdditions(self):
        """ Display pop up menu showing current choice additions """
        viewPopup = Tk()  # pop-up
        viewPopup.wm_title("Choice Additions")

        # TODO
        ttk.Label(viewPopup, text="add some explanation here").pack(side="top", fill="x", pady=10)

        # choice = [[[id, index, label],[...]], ....]
        for choiceList in self.ChoiceRowsToAdd:  # each group of choices
            ttk.Label(viewPopup, text=choiceList[0][0]).pack(side="top", fill="x", pady=10)
            for choice in choiceList:
                ttk.Label(viewPopup, text=choice[2]).pack(side="top", fill="x")
            ttk.Label(viewPopup).pack(side="top", fill="x", pady=10)


    def displayHelpMenu(self):
        """ Display pop up menu explaining survey format """
        helpPopup = Tk() # pop-up

        # Parse about message from textfile
        aboutMessage = open("files/about.txt", "r").readline()
        helpPopup.wm_title("About...")
        ttk.Label(helpPopup, text=aboutMessage, wraplength=400).pack(side="top", fill="x", pady=10)
        helpPopup.mainloop()

    def displayAddNewQuestion(self, qFrame):
        # Form to add new question
        frameQType = Frame(qFrame)
        frameQType.pack(fill=X)

        # Display options for different types of question
        QuestionTypes = {'text', 'decimal', 'integer', 'select_one', 'select_multiple', 'range'}
        Label(frameQType, text="Question Type", width=20).pack(side=LEFT, padx=5, pady=5)
        typeOptionChosen = StringVar(frameQType)   # typeOptionChosen.get() retrieves chosen option
        typeOptionChosen.set('text')                        # default type
        OptionMenu(frameQType, typeOptionChosen, *QuestionTypes).pack(fill=X, padx=5, expand=True)
        # Track type of question chosen and display corresponding question form
        typeOptionChosen.trace("w", lambda name, index, mode,
                                           typeOptionChosen=typeOptionChosen: self.checkTypeOptionChosen(typeOptionChosen, qFrame))

    def clearFrame(self, qFrame):
        # clears frame contents except question_type selector
        flag = False
        for widget in qFrame.winfo_children():
            if (isinstance(widget, Frame)):
                for frame in widget.winfo_children():
                    if qFrame.master is None:   # main question
                        if not(str(frame).startswith('.!frame.')):
                            frame.pack_forget()
                            frame.destroy()
                            widget.destroy()
                    else:                       # follow up question
                        for f in frame.winfo_children():
                            if flag:
                                f.destroy()
                                frame.pack_forget()
                                widget.pack_forget()
                            if (isinstance(f, Menu)):
                                flag = True

    def checkTypeOptionChosen(self, typeOptionChosen, qFrame):
        """ Tracks the typeOptionChosen and displays form accordingly """
        typeOption = typeOptionChosen.get() # type of current question

        self.clearFrame(qFrame)             # clears frame if different option chosen
        questionFrame = Frame(qFrame)
        questionFrame.pack(fill=X, side=TOP)

        questionTypeFrame = Frame(questionFrame)            # questionType
        questionTypeFrame.pack(fill=X)
        Label(questionTypeFrame, text="Question Type", width=20).pack(side=LEFT, padx=5, pady=5)
        ttk.Label(questionTypeFrame, text=typeOption).pack(fill="x", padx=5, pady=5)

        questionNameFrame = Frame(questionFrame)            # questionName
        questionNameFrame.pack(fill=X)
        Label(questionNameFrame, text="Question Name", width=20).pack(side=LEFT, padx=5, pady=5)
        Entry(questionNameFrame).pack(fill=X, padx=5, expand=True)

        questionLabelFrame = Frame(questionFrame)           # questionLabel
        questionLabelFrame.pack(fill=X)
        Label(questionLabelFrame, text="Question Label", width=20).pack(side=LEFT, padx=5, pady=5)
        Entry(questionLabelFrame).pack(fill=X, padx=5, expand=True)

        if (typeOption == 'range'):                     # get range parameters
            self.displayRangeOptions(questionFrame)
        elif (typeOption.startswith('select_')):        # get select_ choice parameters
            self.displaySelectOptions(questionFrame)

        if qFrame.master is None: # main question "Add" button
            # Button to add question
            buttonFrame = Frame(qFrame)
            buttonFrame.pack()
            addButton = Button(buttonFrame, text="Add Question",
                                       command=lambda: self.btnAddQuestion(qFrame, typeOptionChosen))
            addButton.pack(side=BOTTOM, padx=5, pady=5)

    def displayRangeOptions(self, questionFrame):
        """ user must enter "start" "end" "step" parameters for range """
        rangeFrame = Frame(questionFrame)
        rangeFrame.pack(fill=X)
        Label(rangeFrame, text="Enter range parameters", width=20).pack(side=LEFT, padx=5, pady=5)

        # label for start
        startFrame = Frame(questionFrame)
        startFrame.pack(fill=X)
        Label(startFrame, text="Start of range", width=20).pack(side=LEFT, padx=5, pady=5)
        Entry(startFrame).pack(fill=X, padx=5, expand=True)

        # label for end
        endFrame = Frame(questionFrame)
        endFrame.pack(fill=X)
        Label(endFrame, text="End of range", width=20).pack(side=LEFT, padx=5, pady=5)
        Entry(endFrame).pack(fill=X, padx=5, expand=True)

        # label for step
        stepFrame = Frame(questionFrame)
        stepFrame.pack(fill=X)
        Label(stepFrame, text="Step", width=20).pack(side=LEFT, padx=5, pady=5)
        Entry(stepFrame).pack(fill=X, padx=5, expand=True)

    def displaySelectOptions(self, questionFrame):
        # display form for select_ questions allowing user to add in as many choices
        choiceOptionFrame = Frame(questionFrame)
        choiceOptionFrame.pack(fill=X)

        Label(choiceOptionFrame, text="Enter choice options", width=20).pack(side=LEFT, padx=5, pady=5)
        Button(choiceOptionFrame, text="add choice", command=lambda: self.addChoiceDisplay(questionFrame))\
            .pack(side=RIGHT, padx=5, pady=5)

    def addChoiceDisplay(self, questionFrame):
        choiceOptFrame = Frame(questionFrame)           # Add choice options
        choiceOptFrame.pack(fill=X)
        followUpQuestionFrame = Frame(questionFrame)    # Add follow up questions for certain choices
        followUpQuestionFrame.pack(fill=X)

        addFollowUpQuestion = IntVar(choiceOptFrame)
        followUp = Checkbutton(choiceOptFrame, text="Follow Up Question", variable=addFollowUpQuestion, onvalue=1, offvalue=0)
        followUp.pack(side=RIGHT, padx=5, pady=5)
        # Track whether or not follow up chosen for that question
        addFollowUpQuestion.trace("w", lambda name, index, mode,
                                              addFollowUpQuestion=addFollowUpQuestion: self.displayFollowUpQuestion(
            followUpQuestionFrame, addFollowUpQuestion))
        choiceEntry = Entry(choiceOptFrame) # Input choice option label
        choiceEntry.pack(padx=5, pady=5)

    def displayFollowUpQuestion(self, choiceOptFrame, addFollowUpQuestion): # display new frame containing follow up question
            if (addFollowUpQuestion.get()):  # if option selected
                self.displayAddNewQuestion(choiceOptFrame)
            else:
                pass
                # TODO CLEAR CHILD FRAME fix this
                #      # self.clearFrame(choiceOptFrame.master)

    def btnAddQuestion(self, qFrame, typeOptionChosen):
        qType = typeOptionChosen.get()  # parse question according to their type
        self.parseQuestionForm(qFrame, str(qType))
        self.clearFrame(qFrame)         # resets frame

    def parseQuestionForm(self, qFrame, qType):
        qFormEntries   = []  # get text entries from form
        followupFrames = []  # get follow-up questions from form

        for widget in qFrame.winfo_children():  # contents of entry boxes
            for child_widget in widget.winfo_children():
                for cchild_widget in child_widget.winfo_children():
                    if isinstance(cchild_widget,Entry):  # top level text entry
                        qFormEntries.append(cchild_widget.get())
                    elif isinstance(cchild_widget, Frame):  # follow-up question
                        followupFrames.append(cchild_widget)


        if (qType == 'range'):
            self.parseRangeQuestion(qType, qFormEntries)                      # parse range question
        elif (str(qType).startswith('select_')):
            self.parseSelectQuestion(qType, qFormEntries, followupFrames)     # parse select_ question
        else:
            self.parseInputQuestion(qType, qFormEntries)                      # parse input question

    def parseRangeQuestion(self, qType, rangeEntry):
        # rangeEntry = [<qName>, <qLabel>, <start>, <step>, <end>]
        parameters = "start="+rangeEntry[2] + " end="+rangeEntry[3] + " step="+rangeEntry[4]

        # question = [qType, qName, qLabel, relevant, parameters, media]
        self.addQuestionRow(qType, rangeEntry[0], rangeEntry[1], '', parameters, '')

    def parseInputQuestion(self, qType, inputEntry):
        # inputEntry = [<qName>, <qLabel>]
        self.addQuestionRow(qType, inputEntry[0], inputEntry[1], '', '', '')

    def parseSelectQuestion(self, qType, selectEntry, followupFrames):
        # selectEntry = [<qName>, <qLabel>, <choiceopt1> , <choiceopt2> ....]
        qName = selectEntry[0]
        qLabel = selectEntry[1]

        choiceEntries = selectEntry[2:]     # rest of entries are choice options

        selectType = qType + " " + qName    # e.g. select_one gender
        self.addQuestionRow(selectType, qName, qLabel, '', '', '')
        self.addChoicesRows(qName, choiceEntries)

        currentChoice = 1
        for frame in followupFrames:
            for widget in frame.winfo_children():
                for w in widget.winfo_children():
                    if (str(w).endswith('!label2')):
                        followupType = (w.cget("text"))                # get type of follow-up question
                        relevance = "{%s}=%d" % (qName, currentChoice) # e.g {gender}=1
                        self.addQuestionRow('begin_group','','', relevance,'','')
                        self.parseQuestionForm(frame.master, followupType)
                        self.addQuestionRow('end_group', '', '', '', '', '')
                        currentChoice += 1

    def addQuestionRow(self, qType, qName, qLabel, relevant, parameters, media):
        # add new question to self.QuestionRowsToAdd
        rowValues = [qType, qName, qLabel, relevant, parameters, media]
        self.QuestionRowsToAdd.append(rowValues)

    def addChoicesRows(self, qName, choiceOptions):
        # add new list of choice options to self.ChoiceRowsToAdd
        choiceOptionsRows = [] # = [[listName, name, label], [..], ....]

        index = 1
        for choice in choiceOptions:
            choiceOptionsRows.append([qName, index, choice])
            index += 1
        self.ChoiceRowsToAdd.append(choiceOptionsRows)

    def save(self):
        # # save changes to workbook
        # filename = 'files/survey-choices-update.xlsx'
        # wb = load_workbook(filename)
        #
        # # surveySheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        #
        # # add choice options to choices sheet
        # choiceSheet = wb.get_sheet_by_name(wb.get_sheet_names()[1])
        # nRows = choiceSheet.max_row # index of last row
        # nRows += 2 # leave a gap between different lists
        #
        # for choice_list in self.ChoiceRowsToAdd:
        #     for choice_row in choice_list:
        #         c = 1
        #         for row_cell in choice_row:
        #             choiceSheet.cell(row=nRows, column=c).value = row_cell
        #             c += 1
        #         nRows += 1
        #     nRows += 1  # row gap between adjacent lists
        #
        # # save changes
        # wb.save(filename)

        print("Questions")
        print(self.QuestionRowsToAdd)
        print("Choices")
        print(self.ChoiceRowsToAdd)
        # # Load in the workbook
        # wb = load_workbook(filename)
        # surveySheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        #
        # # number of rows/columns in original worksheet
        # nRows = surveySheet.max_row
        # nCols = surveySheet.max_column
        #
        # for index in range(0, nCols):
        #     surveySheet.cell(row=nRows + 1, column=index + 1).value = rowValues[index]
        #
        # # save changes
        # wb.save(filename)

        # clear current cache
        self.QuestionRowsToAdd = []
        self.ChoiceRowsToAdd = []

    def onExit(self):
        self.quit() # close window




def main():
    root = Tk()
    # root.geometry("250x200+300+300")
    app = Example()
    root.mainloop()

if __name__ == '__main__':
    main()

# todo by default add all before end section

# todo seperate question names by hyphens if not already
# todo some way of indenting thats more understandable or colours??

