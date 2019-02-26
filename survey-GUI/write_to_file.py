from openpyxl import load_workbook


def writeQuestions(filename, question_rows):
    wb = load_workbook(filename)

    # add questions to survey sheet
    survey_sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    n_rows = survey_sheet.max_row + 1  # start on next empty row
    n_cols = survey_sheet.max_column

    for q_row in question_rows:
        for index in range(0, n_cols):
            survey_sheet.cell(row=n_rows, column=index + 1).value = q_row[index]
        n_rows += 1

    # save changes to sheet
    wb.save(filename)


def writeChoices(filename, choice_rows):
    wb = load_workbook(filename)
    # add choice options to choices sheet
    choice_sheet = wb.get_sheet_by_name(wb.get_sheet_names()[1])

    n_rows = choice_sheet.max_row  # index of last row
    n_rows += 2  # leave a gap between different lists

    for choice_list in choice_rows:
        for choice_row in choice_list:
            c = 1
            for row_cell in choice_row:
                choice_sheet.cell(row=n_rows, column=c).value = row_cell
                c += 1
            n_rows += 1
        n_rows += 1  # row gap between adjacent lists

    # save changes to sheet
    wb.save(filename)
